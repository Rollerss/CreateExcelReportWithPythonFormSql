import openpyxl
import pyodbc


def create_excel_from_sql_data(columnNames, sql_data, excel_file_name):
    # Create a workbook and add a worksheet.
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SQL Data"

    # Add the headers
    for i in range(len(columnNames)):
        worksheet.cell(row=1, column=i + 1).value = columnNames[i]

    # Add some data, we can use the range() function to add a list of values.
    for row in range(1, len(sql_data) + 1):
        for col in range(1, len(sql_data[0]) + 1):
            worksheet.cell(row=row + 1, column=col).value = sql_data[row - 1][col - 1]

    # Save the workbook.
    workbook.save(excel_file_name)


def get_sql_data(query, connStr):
    conn = pyodbc.connect(connStr)
    cursor = conn.cursor()
    cursor.execute(query)
    return cursor.fetchall()


def create_query(tableName, columnNames):
    query = "SELECT "
    for columnName in columnNames:
        query += "[" + columnName + "], "
    query = query[:-2]
    query += " FROM " + tableName
    return query


def create_connection_string(server, database):
    return (
        "DRIVER={ODBC Driver 17 for SQL Server};SERVER="
        + f"{server};DATABASE={database};Trusted_Connection=yes;"
    )


def create_customer_report(folderName):
    columnNames = ["CustomerId", "Name", "Location", "Email"]
    server = "(localdb)\MSSQLLocalDB"
    database = "TutorialDB"
    tableName = "Customers"
    fileName = f"{folderName}SQLData2.xlsx"

    connStr = create_connection_string(server, database)
    query = create_query(tableName, columnNames)
    sql_data = get_sql_data(query, connStr)
    create_excel_from_sql_data(columnNames, sql_data, fileName)


def main():
    folderName = "C:\\TestFiles\\"
    print("start")
    create_customer_report(folderName)
    print("end")


if __name__ == "__main__":
    main()
